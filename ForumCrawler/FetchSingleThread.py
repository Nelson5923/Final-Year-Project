import requests
from lxml import etree
from selenium import webdriver
import json
import re
import pandas as pd
from pandas import ExcelWriter
import time
import openpyxl
from openpyxl import load_workbook
import os.path
import sqlite3

def getPageNum(url, driver):

    # Get the Post
    driver.get(url)
    Page = driver.page_source
    Tree = etree.HTML(Page)
    DataNode = ''.join(Tree.xpath('//script/text()'))
    Data = re.search(r'window.__PRELOADED_STATE__ = {.*}$', DataNode).group()
    ParsedData = ''.join(re.search(r'{.*}', Data).group())

    # Process the JSON data
    j = json.loads(ParsedData)  # print(j['topic']) # print(j['thread']) # print(j['thread']['replies'])

    return j['thread']['totalPage']

def fetch(url,PageNum,driver, ThreadID):

    Columns = ['ReplyID', 'Forum', 'Index', 'AuthorID', 'AuthorName', 'AuthorGender', 'Content', 'ReplyDate', 'ThreadID']
    Thread = pd.DataFrame(columns=Columns, dtype=object)
    Database = sqlite3.connect('Thread.sqlite3')

    for i in range(1,PageNum+1):

        # Get the Post
        print("Getting:" + url + "&page=" + str(i))
        driver.get(url + "&page=" + str(i))
        time.sleep(2)

        Page = driver.page_source
        Tree = etree.HTML(Page)
        DataNode = ''.join(Tree.xpath('//script/text()'))
        Data = re.search(r'window.__PRELOADED_STATE__ = {.*}$', DataNode).group()
        ParsedData = ''.join(re.search(r'{.*}', Data).group())

        # Process the JSON data
        j = json.loads(ParsedData)  # print(j['topic']) # print(j['thread']) # print(j['thread']['replies'])
        for Post in j['thread']['replies']:
            Post['content'] = re.sub('&gt;', '>', Post['content'])
            Post['content'] = re.sub('&quot;', '\\"', Post['content'])
            Post['content'] = re.sub('\r<br />', '\n', Post['content'])
            Post['content'] = re.sub('<br />', '\n', Post['content'])
            Post['content'] = "<body>" + Post['content'] + "</body>"
            Tree = etree.HTML(Post['content'])
            Post['content'] = ''.join(Tree.xpath('//body/text()'))
            Post['content'] = re.sub('<.*?>', '', Post['content'])
            Post['content'] = re.sub('[;\n\\.,，。]', ' ', Post['content'])
            Post['content'] = re.sub(' {2,}', ' ', Post['content'])
            Post['content'] = Post['content'].strip(' ');

        for Post in j['thread']['replies']:
            if len(Post['content']) > 1:
                Post['ThreadID'] = ThreadID
                Record = pd.DataFrame([list(Post.values())], columns=Columns, dtype=object)
                ExportToSQLite(Database, Record)
                Thread = Thread.append(Record)

    Thread.set_index('Index', inplace=True)
    Thread.reset_index(drop=True, inplace=True)
    Thread.index.name = 'Index'

    print(Thread.to_string())
    print("Complete" + ":" + url)

    return Thread

def ExportToSQLite(Database, Record):
    try:
        Record.to_sql('Thread', Database, index=False, if_exists='append')
    except sqlite3.IntegrityError:
        pass

def ExportToCSV(Thread):
    if os.path.isfile('Thread.csv'):
        Thread.to_csv('Thread.csv', encoding='utf_8_sig', mode='a', header=False)
    else:
        Thread.to_csv('Thread.csv', encoding='utf_8_sig')

def ExportToExcel(Thread):  # Not in Used

    writer = ExcelWriter('Thread.xlsx')

    if os.path.isfile('Thread.xlsx'):
        book = load_workbook('Thread.xlsx')
        if ThreadID in book.sheetnames:
            stn = book.get_sheet_by_name(ThreadID)
            book.remove_sheet(stn)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    Thread.to_excel(writer, 'Sheet')
    writer.save()

def main():

    # Disable Browser Pop up
    Option = webdriver.ChromeOptions()
    Option.add_argument('headless')

    # Disable Image
    prefs = {"profile.managed_default_content_settings.images": 2}
    Option.add_experimental_option("prefs", prefs)

    # Get the Javascript Website
    driver = webdriver.Chrome(executable_path='C:\Chrome\chromedriver.exe', chrome_options=Option)

    # Fetch
    url = "https://hkug.arukascloud.io/topics/2/6975952?forum=HKG"
    PageNum = getPageNum(url, driver)
    Thread = fetch(url,PageNum,driver,str(873403))

    # Closed the Driver
    driver.close();

if __name__ == '__main__': # Function inside this block will not be Executed
    main();